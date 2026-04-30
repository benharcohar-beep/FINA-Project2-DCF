# =============================================================================
# DCF Equity Valuation — Streamlit App
# =============================================================================
# Discounted Cash Flow valuation tool with optional SEC EDGAR auto-fill,
# CAPM-based WACC build-up, multi-stage growth, margin expansion path,
# sensitivity tables, and Excel export with live formulas.
#
# Run locally:
#   pip install -r requirements.txt
#   streamlit run equity_valuation_app.py
# =============================================================================

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import requests
from datetime import datetime
from io import BytesIO

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DCF Equity Valuation",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.title("📊 DCF Equity Valuation")
st.caption("Discounted Cash Flow valuation tool with SEC filings auto-fill and Excel export")

# =============================================================================
# SEC EDGAR — optional auto-fill helpers
# Uses the official SEC API: free, no key, not rate-limited beyond UA requirement.
# =============================================================================
_SEC_HEADERS = {
    "User-Agent": "FINA Project2 DCF Valuation benharcohar@gmail.com",
    "Accept-Encoding": "gzip, deflate",
}

# Sentinel returned by EDGAR helpers to distinguish "the SEC API blocked / failed
# us right now" (transient, retryable) from "this ticker truly isn't in EDGAR"
# (permanent for this symbol). The app surfaces different messages for each.
_FETCH_FAILED = object()

@st.cache_data(ttl=86400, show_spinner=False)
def _ticker_to_cik_cached(symbol: str):
    """Cached path — only stores successful (cik, name) tuples."""
    sym_u = symbol.upper().strip()
    variants = {sym_u, sym_u.replace(".", "-"), sym_u.replace("-", "."), sym_u.replace(".", "")}
    r = requests.get(
        "https://www.sec.gov/files/company_tickers.json",
        headers={"User-Agent": _SEC_HEADERS["User-Agent"]},
        timeout=10,
    )
    r.raise_for_status()
    for entry in r.json().values():
        if entry.get("ticker", "").upper() in variants:
            return str(entry["cik_str"]).zfill(10), entry.get("title", symbol)
    return None, None  # ticker genuinely not in SEC index → safe to cache


def _ticker_to_cik(symbol: str):
    """Map ticker → 10-digit CIK + company name.

    Returns:
        (cik, name)  on success
        (None, None) when the ticker isn't in EDGAR's index (cached for a day)
        _FETCH_FAILED when the SEC blocked/timed out (NOT cached, retry on next call)
    """
    if not symbol:
        return None, None
    try:
        return _ticker_to_cik_cached(symbol)
    except Exception:
        # Transient failure — the @st.cache_data decorator does not cache exceptions,
        # so the next call will retry instead of getting a stale "not found".
        return _FETCH_FAILED


def _latest_annual(facts_block):
    """Extract the most recent annual value from an XBRL concept block.

    EDGAR rows have multiple periods (quarterly, YTD, full-year). We detect
    annual values by duration (start→end ≈ 365 days) and group by the year of
    the `end` date. Works for any fiscal calendar.
    """
    if not facts_block:
        return None, []
    rows = facts_block.get("units", {}).get("USD") \
        or facts_block.get("units", {}).get("shares") \
        or next(iter(facts_block.get("units", {}).values()), [])

    annual = []
    for r in rows:
        start, end = r.get("start"), r.get("end")
        if not end:
            continue
        # Instant facts (balance-sheet line items) have no `start`
        if not start:
            annual.append(r)
            continue
        try:
            d = (datetime.fromisoformat(end) - datetime.fromisoformat(start)).days
            if 350 <= d <= 380:
                annual.append(r)
        except Exception:
            continue

    # Prefer 10-K filings
    tenk = [r for r in annual if r.get("form") == "10-K"]
    if tenk:
        annual = tenk

    # Group by year-of-end, keep latest filed
    by_year = {}
    for r in annual:
        try:
            yk = datetime.fromisoformat(r["end"]).year
        except Exception:
            continue
        if yk not in by_year or r.get("filed", "") > by_year[yk].get("filed", ""):
            by_year[yk] = r

    series = sorted(by_year.items(), key=lambda kv: kv[0], reverse=True)
    latest = series[0][1]["val"] if series else None
    return latest, [(y, row["val"]) for y, row in series if row.get("val") is not None]


def _best_concept(us_gaap, *names):
    """Return the concept block with the freshest annual data among the given tags.
    Companies switch tags over time (e.g. ASC 606 moved revenue from `Revenues`
    to `RevenueFromContractWithCustomerExcludingAssessedTax`).
    """
    best, best_year = None, -1
    for n in names:
        block = us_gaap.get(n)
        if not block:
            continue
        rows = block.get("units", {}).get("USD") \
            or block.get("units", {}).get("shares") or []
        years = [r.get("fy") for r in rows if r.get("fp") == "FY" and r.get("fy")]
        if not years:
            continue
        max_fy = max(years)
        if max_fy > best_year:
            best, best_year = block, max_fy
    return best


@st.cache_data(ttl=3600, show_spinner=False)
def _fetch_edgar_cached(symbol: str):
    """Cached path — only stores successful dicts and definitive 'not found' (None)."""
    cik, name = _ticker_to_cik_cached(symbol)
    if cik is None:
        return None  # genuinely unknown ticker — OK to cache
    r = requests.get(
        f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json",
        headers={**_SEC_HEADERS, "Host": "data.sec.gov"},
        timeout=15,
    )
    r.raise_for_status()  # any non-200 → exception → not cached, will retry
    ug = r.json().get("facts", {}).get("us-gaap", {}) or {}
    dei = r.json().get("facts", {}).get("dei", {}) or {}
    # Combined namespace lookup for shares (dei + us-gaap — companies use either)
    all_facts = {**ug, **dei}

    rev,    rev_hist = _latest_annual(_best_concept(ug,
        "Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax",
        "SalesRevenueNet", "RevenueFromContractWithCustomerIncludingAssessedTax"))
    # Operating income — use NetIncomeLoss as fallback for banks/insurers that don't report OpIncLoss
    opinc,  _        = _latest_annual(_best_concept(ug,
        "OperatingIncomeLoss", "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest"))
    net_inc, _       = _latest_annual(_best_concept(ug, "NetIncomeLoss",
        "ProfitLoss", "NetIncomeLossAvailableToCommonStockholdersBasic"))
    ocf,    _        = _latest_annual(_best_concept(ug,
        "NetCashProvidedByUsedInOperatingActivities"))
    capex,  _        = _latest_annual(_best_concept(ug,
        "PaymentsToAcquirePropertyPlantAndEquipment",
        "PaymentsToAcquireProductiveAssets"))
    cash,   _        = _latest_annual(_best_concept(ug,
        "CashAndCashEquivalentsAtCarryingValue",
        "Cash",
        "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"))
    dlt,    _        = _latest_annual(_best_concept(ug,
        "LongTermDebt", "LongTermDebtNoncurrent"))
    dst,    _        = _latest_annual(_best_concept(ug,
        "LongTermDebtCurrent", "DebtCurrent",
        "ShortTermBorrowings"))
    # Shares: try multiple tags across BOTH dei and us-gaap. Prefer diluted weighted-avg
    # (standard for per-share metrics), fall back to common shares outstanding.
    shares, _        = _latest_annual(_best_concept(all_facts,
        "WeightedAverageNumberOfDilutedSharesOutstanding",
        "WeightedAverageNumberOfSharesOutstandingBasic",
        "CommonStockSharesOutstanding",
        "EntityCommonStockSharesOutstanding"))

    # EBIT margin: prefer real EBIT, else estimate from net income (assume 21% effective tax)
    ebit_margin = None
    if opinc and rev and opinc > 0:
        ebit_margin = opinc / rev
    elif net_inc and rev and net_inc > 0:
        ebit_margin = (net_inc / 0.79) / rev

    revenue_cagr = None
    if rev_hist and len(rev_hist) >= 2:
        n_years = min(len(rev_hist) - 1, 4)
        oldest = rev_hist[n_years][1]
        latest = rev_hist[0][1]
        if oldest > 0 and n_years > 0:
            revenue_cagr = (latest / oldest) ** (1 / n_years) - 1

    missing = []
    if not rev:        missing.append("Revenue")
    if not shares:     missing.append("Shares Outstanding")
    if ebit_margin is None: missing.append("EBIT Margin")
    if cash is None:   missing.append("Cash")
    if not (dlt or dst): missing.append("Debt")

    return {
        "name":          name,
        "revenue":       rev / 1e6 if rev else None,
        "ebit_margin":   ebit_margin * 100 if ebit_margin else None,
        "operating_cf":  ocf / 1e6 if ocf else None,
        "capex":         capex / 1e6 if capex else None,
        "cash":          cash / 1e6 if cash else None,
        "debt":          ((dlt or 0) + (dst or 0)) / 1e6 if (dlt or dst) else None,
        "shares":        shares / 1e6 if shares else None,
        "revenue_cagr":  revenue_cagr * 100 if revenue_cagr else None,
        "rev_history":   [(y, v / 1e9) for y, v in rev_hist[:5]],
        "fy_end":        rev_hist[0][0] if rev_hist else None,
        "missing":       missing,
    }


def fetch_edgar_fundamentals(symbol: str):
    """Pull fundamentals from SEC EDGAR Company Facts API.

    Returns:
        dict          — fundamentals on success
        None          — ticker confirmed not in EDGAR (cached for 1h)
        _FETCH_FAILED — SEC API blocked/timed out (NOT cached, retry on next call)
    """
    if not symbol:
        return None
    try:
        return _fetch_edgar_cached(symbol)
    except Exception:
        return _FETCH_FAILED


# =============================================================================
# Session-state defaults — every input has a fallback so the app always runs
# =============================================================================
DEFAULTS = {
    "ticker":          "AAPL",
    "company_name":    "Apple Inc.",
    "revenue":         391000.0,   # $M (AAPL FY24 starting point)
    "growth":          8.0,        # % — modest growth above GDP
    "margin":          30.0,       # % EBIT margin
    "tax_rate":        21.0,       # %
    "reinvest_rate":   15.0,       # % of NOPAT — mature firms reinvest less
    "wacc":            8.5,        # % — large-cap blue chip
    "terminal_growth": 2.5,        # %
    "years":           5,
    "debt":            107000.0,   # $M
    "cash":            65000.0,    # $M
    "shares":          15300.0,    # millions
    "current_price":   210.00,     # $
    "edgar_data":      None,
    "loaded_ticker":   "AAPL",
    "edgar_msg":       None,
    # ── Advanced modeling toggles ────────────────────────────────────────
    "wacc_mode":       "Direct",   # "Direct" | "CAPM build-up"
    "rf":              4.00,       # Risk-free rate (10-yr Treasury), %
    "erp":             5.50,       # Equity risk premium, %
    "beta":            1.00,       # Levered beta
    "rd_pretax":       5.00,       # Pre-tax cost of debt, %
    "weight_debt":     20.0,       # Target debt weight, % (rest = equity)
    "growth_mode":     "Constant", # "Constant" | "3-stage (high → fade → terminal)"
    "high_growth":     15.0,
    "high_years":      3,
    "fade_years":      4,
    "margin_mode":     "Constant", # "Constant" | "Linear expansion"
    "margin_terminal": 30.0,
    # Terminal value method
    "tv_method":       "Gordon Growth", # "Gordon Growth" | "Exit Multiple"
    "exit_multiple":   12.0,       # EV / EBITDA at year N
    # Discounting convention
    "mid_year":        False,      # mid-year convention (more accurate)
    # Verdict thresholds
    "fair_band":       10.0,       # % band around price that counts as "fairly valued"
    # Working capital change as a separate driver
    "nwc_pct":         2.0,        # change in NWC as % of revenue (deducted from FCF)
}
# Process pending reset BEFORE any widget instantiates — Streamlit doesn't
# allow modifying a widget's session_state after the widget has been rendered
# in the same script run. Reset button just sets a flag; the actual restore
# happens here at the top of the next run.
if st.session_state.pop("_pending_reset", False):
    for k in list(st.session_state.keys()):
        if k in DEFAULTS:
            del st.session_state[k]

# Initialize defaults for any keys not yet in session_state
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


def apply_edgar_autofill():
    """Fetch EDGAR data for the current ticker and apply it to all inputs.

    Used by both the explicit "Auto-fill" button AND the ticker on_change callback,
    so changing the ticker (Tab/Enter) immediately refreshes the model.
    Sets st.session_state.edgar_msg so the sidebar can render the result inline.
    """
    sym = (st.session_state.ticker or "").strip().upper()
    if not sym:
        return
    edgar = fetch_edgar_fundamentals(sym)
    if edgar is _FETCH_FAILED:
        st.session_state.edgar_msg = ("error",
            f"⚠️ Couldn't reach SEC EDGAR right now (server returned an error or timed out). "
            f"This is usually transient — wait a few seconds and click **🔄 Refresh** to retry. "
            f"You can keep editing manual inputs below in the meantime.")
        return
    if edgar is None:
        st.session_state.edgar_msg = ("error",
            f"❌ Ticker **{sym}** not found in EDGAR. Use a US-listed ticker, "
            "or enter values manually below.")
        return
    edgar_rev = edgar.get("revenue")
    if not edgar_rev or edgar_rev <= 0:
        st.session_state.edgar_data = None
        st.session_state.edgar_msg = ("error",
            f"⛔ **{edgar.get('name', sym)}** appears to be a **pre-revenue company** "
            "(no operating revenue in latest 10-K). DCF requires positive cash flows. "
            "Try a revenue-generating ticker (AAPL, MSFT, JNJ), or use asset-based valuation, "
            "or enter values manually below.")
        return

    # Slider bounds for keys that are bound to sliders. If EDGAR returns a value
    # outside these bounds (e.g. NVDA's ~68% revenue CAGR vs. growth slider max),
    # writing it raw into session_state crashes Streamlit when the slider re-renders.
    # Clamp here so the app stays alive; the user can still adjust manually.
    SLIDER_BOUNDS = {
        "growth": (-20.0, 100.0),
        "margin": (0.0, 100.0),
    }

    def _set(key, val, min_val=0):
        if val is None: return
        try: v = float(val)
        except (TypeError, ValueError): return
        if v <= min_val: return
        if key in SLIDER_BOUNDS:
            lo, hi = SLIDER_BOUNDS[key]
            v = max(lo, min(hi, v))
        st.session_state[key] = round(v, 1) if isinstance(v, float) and abs(v) < 1000 else round(v)

    if edgar.get("name"):
        st.session_state.company_name = edgar["name"]
    _set("revenue", edgar.get("revenue"))
    _set("margin",  edgar.get("ebit_margin"))
    _set("shares",  edgar.get("shares"))
    _set("growth",  edgar.get("revenue_cagr"))
    if edgar.get("cash") is not None and edgar.get("cash") >= 0:
        st.session_state.cash = round(edgar["cash"], 0)
    if edgar.get("debt") is not None and edgar.get("debt") >= 0:
        st.session_state.debt = round(edgar["debt"], 0)
    if edgar.get("capex") and edgar.get("revenue") and edgar.get("ebit_margin"):
        nopat = edgar["revenue"] * (edgar["ebit_margin"]/100) * (1 - st.session_state.tax_rate/100)
        if nopat > 0:
            ri = min(100, edgar["capex"] / nopat * 100)
            if ri > 0:
                st.session_state.reinvest_rate = round(ri, 1)
    st.session_state.edgar_data = edgar
    st.session_state.loaded_ticker = sym
    fy = edgar.get("fy_end", "?")
    if edgar.get("ebit_margin") is None:
        st.session_state.edgar_msg = ("warning",
            f"⚠️ Loaded {edgar.get('name', sym)} (FY{fy}), but operating margin "
            "couldn't be determined (likely a loss year). DCF will reflect your manual "
            "margin assumption, not the company's actuals.")
    elif edgar.get("missing"):
        st.session_state.edgar_msg = ("warning",
            f"⚠️ Loaded {edgar.get('name', sym)} (FY{fy}). EDGAR didn't have: "
            f"**{', '.join(edgar['missing'])}** — left at default. Edit fields below if needed.")
    else:
        st.session_state.edgar_msg = ("success",
            f"✅ Loaded {edgar.get('name', sym)} (FY{fy})")


def on_ticker_change():
    """Streamlit callback fired when the ticker text_input commits.
    Only auto-fetch if the ticker actually changed since last load — avoids
    a wasted EDGAR call when the user just tabs through the field.
    """
    cur = (st.session_state.get("ticker") or "").strip().upper()
    loaded = (st.session_state.get("loaded_ticker") or "").strip().upper()
    if cur and cur != loaded:
        apply_edgar_autofill()


# =============================================================================
# Sidebar — all inputs
# =============================================================================
with st.sidebar:
    st.header("📋 Inputs")

    # ─── Ticker + EDGAR auto-fill ────────────────────────────────────────────
    st.subheader("🏢 Company")
    st.text_input(
        "Ticker", key="ticker",
        on_change=on_ticker_change,
        help="US-listed ticker (e.g. AAPL, MSFT, JNJ). Press Enter or Tab to auto-load EDGAR data.",
    )

    col_a, col_b = st.columns([3, 2])
    with col_a:
        if st.button("🔄 Re-fetch from EDGAR", use_container_width=True,
                     help="Re-pull latest 10-K data from SEC EDGAR for the current ticker"):
            apply_edgar_autofill()
            st.rerun()
    with col_b:
        if st.button("🧹 Reset", use_container_width=True, help="Reset all inputs to defaults"):
            # Defer the actual reset to the top of the next run — see notes there
            st.session_state["_pending_reset"] = True
            st.session_state["edgar_msg"] = None
            st.rerun()

    # ── Render the last EDGAR-fetch message inline (set by the callback / button) ──
    msg = st.session_state.get("edgar_msg")
    if msg:
        kind, text = msg
        {"success": st.success, "warning": st.warning, "error": st.error}[kind](text)

    # ── Mismatch banner: ticker changed but data is stale (network failed?) ──
    cur = (st.session_state.ticker or "").strip().upper()
    loaded = (st.session_state.get("loaded_ticker") or "").strip().upper()
    if cur and loaded and cur != loaded:
        st.info(
            f"📌 Inputs below currently reflect **{loaded}**. "
            f"Press Enter or click 'Re-fetch from EDGAR' to load **{cur}**."
        )

    # Company Name and Current Revenue widgets intentionally hidden — the
    # values are populated by EDGAR autofill (or DEFAULTS) and flow into the
    # DCF math from session_state. The customer doesn't see the raw values.

    st.markdown("---")

    # ─── Financial inputs ────────────────────────────────────────────────────
    st.subheader("💰 Financial Inputs")

    # ── Margin: simple slider, with optional linear-expansion mode ──
    st.radio("Margin assumption", options=["Constant", "Linear expansion"],
             key="margin_mode", horizontal=True,
             help="Constant = flat EBIT margin every year. Linear expansion = margin moves linearly from Year 1 to Year N (good for growth firms with operating leverage).")
    st.slider("EBIT Margin (Year 1) (%)", min_value=0.0, max_value=100.0, step=0.5, key="margin",
              format="%.1f%%",
              help="Operating Income ÷ Revenue. Higher margin → more value. EDGAR auto-fills from 10-K.\n\n"
                  "**Recommended:** depends heavily on industry. Software/tech: 25-40%. Pharma: 20-30%. "
                  "Banks/insurers: 25-35%. Consumer staples: 15-25%. Retail/grocers: 3-8%. Airlines/auto: 5-12%. "
                  "Use the EDGAR-loaded value as your baseline.")
    if st.session_state.margin_mode == "Linear expansion":
        st.slider("Terminal EBIT Margin (Year N) (%)", min_value=0.0, max_value=100.0, step=0.5,
                  key="margin_terminal", format="%.1f%%",
                  help="Margin in the final forecast year. Linear interpolation between Year 1 and Year N.\n\n"
                       "**Recommended:** for growth firms with operating leverage, terminal margin is typically "
                       "3-10 percentage points above Year 1. For mature firms, set roughly equal to Year 1.")

    st.slider("Tax Rate (%)", min_value=0.0, max_value=50.0, step=0.5, key="tax_rate",
              format="%.1f%%",
              help="Effective corporate tax rate. US federal statutory is 21%; effective often 15-25%.\n\n"
                   "**Recommended:** **21%** for US federal-only. **25%** for US federal + average state tax. "
                   "Most large-cap US filers run **18-22%** effective. Multinationals with foreign-income shields "
                   "(AAPL, MSFT) often **13-18%**. Use the company's last 10-K effective rate as a baseline.")
    st.slider("Reinvestment Rate (%)", min_value=0.0, max_value=100.0, step=1.0, key="reinvest_rate",
              format="%.0f%%",
              help="% of NOPAT reinvested in CapEx + working capital. Higher → less FCF today, more future growth.\n\n"
                   "**Recommended:** mature/cash-cow firms (AAPL, KO, JNJ): **10-25%**. "
                   "Steady growth: **25-40%**. High growth/capital-heavy (TSLA, NVDA, semis, telecoms): **40-70%**. "
                   "Rule of thumb: reinvestment ≈ growth ÷ ROIC.")

    st.slider("Δ Working Capital (% of Δ Revenue)", min_value=0.0, max_value=30.0, step=0.5,
              key="nwc_pct", format="%.1f%%",
              help="Net working capital tied up per dollar of revenue growth. Subtracted from FCF.\n\n"
                   "**Recommended:** **2%** as a default. Software / asset-light: **0-2%** (negative deferred-revenue "
                   "tailwind possible). Manufacturing / industrial: **3-7%**. Retail / inventory-heavy: **5-10%**.")

    st.markdown("---")

    # ─── Growth ──────────────────────────────────────────────────────────────
    st.subheader("📈 Growth")
    # Surface historical growth from EDGAR right next to the growth slider — gives the
    # user a concrete reference point to anchor their forecast.
    if st.session_state.edgar_data:
        e = st.session_state.edgar_data
        if e.get("rev_history") and len(e["rev_history"]) >= 2:
            yoy = []
            for i in range(min(3, len(e["rev_history"]) - 1)):
                cur = e["rev_history"][i][1]; prev = e["rev_history"][i+1][1]
                if prev > 0: yoy.append((cur/prev - 1) * 100)
            if yoy:
                yoy_str = " / ".join(f"{g:+.1f}%" for g in yoy)
                cagr = e.get("revenue_cagr")
                cagr_str = f" · 3-yr CAGR: **{cagr:.1f}%**" if cagr else ""
                st.caption(f"📊 Historical YoY (most recent first): {yoy_str}{cagr_str}")

    st.radio("Growth path", options=["Constant", "3-stage (high → fade → terminal)"],
             key="growth_mode", horizontal=False,
             help="Constant = same growth rate every year. 3-stage = high-growth period, then linear fade, then stable terminal.")

    if st.session_state.growth_mode == "Constant":
        st.slider("Annual Growth Rate (%)", min_value=-20.0, max_value=100.0, step=0.5, key="growth",
                  format="%.1f%%",
                  help="Revenue growth during the forecast period. EDGAR fills with 3-yr CAGR.\n\n"
                       "**Recommended:** mature large-caps: **3-7%**. Above-GDP growth firms: **8-15%**. "
                       "High-growth: **15-25%**. Hypergrowth (NVDA, SHOP): **25%+** but switch to **3-stage** mode — "
                       "perpetual high growth is rarely realistic.")
    else:
        st.slider("High-growth Rate (%)", min_value=-10.0, max_value=80.0, step=0.5,
                  key="high_growth", format="%.1f%%",
                  help="Growth rate during the high-growth phase.\n\n"
                       "**Recommended:** strong large-caps: **10-15%**. Growth companies: **15-25%**. "
                       "Hypergrowth (early-stage tech, semis): **25-50%**. Match this to recent reported revenue growth.")
        st.slider("High-growth Years", min_value=1, max_value=10,
                  key="high_years",
                  help="Number of years at the high-growth rate before fading begins.\n\n"
                       "**Recommended:** **3 years** for most growth firms. **5 years** for companies with strong "
                       "competitive moats or large addressable markets. Beyond 5-7 years, fade is more honest.")
        st.slider("Fade-down Years", min_value=1, max_value=10,
                  key="fade_years",
                  help="Number of years over which growth linearly fades from high-growth rate to terminal rate.\n\n"
                       "**Recommended:** **4-7 years**. Longer fades imply slower competitive erosion — appropriate "
                       "for moat-protected firms. Shorter fades (2-3 yrs) for commodity-like or cyclical businesses.")
        # In 3-stage mode, the simple `growth` slider is unused — show it disabled for transparency
        st.caption("(The simple 'Annual Growth Rate' slider is ignored in 3-stage mode.)")

    st.slider("Terminal Growth Rate (%)", min_value=0.0, max_value=10.0, step=0.25,
              key="terminal_growth", format="%.2f%%",
              help="Perpetual growth after the forecast period.\n\n"
                   "**Recommended:** **2.5%** as a default (long-run US real GDP growth + inflation target). "
                   "Hard ceiling: **3%**. Going above implies the company eventually grows faster than the whole economy — "
                   "mathematically impossible long-term. Below 2% for declining industries.")

    st.slider("Forecast Years", min_value=3, max_value=15, key="years",
              help="Length of the explicit forecast.\n\n"
                   "**Recommended:** **5 years** for stable mature firms. **7-10 years** for growth firms where "
                   "near-term cash flows differ materially from steady-state. Banks/research notes typically use 5; "
                   "academic / textbook DCFs typically use 10.")

    # ── Terminal Value method ──
    st.radio("Terminal Value method", options=["Gordon Growth", "Exit Multiple"],
             key="tv_method", horizontal=True,
             help="Gordon Growth = perpetuity formula using terminal growth rate. Exit Multiple = EV/EBITDA × terminal-year EBITDA (assumes the firm is sold at year N).")
    if st.session_state.tv_method == "Exit Multiple":
        st.slider("Exit EV/EBITDA Multiple", min_value=2.0, max_value=40.0, step=0.5,
                  key="exit_multiple", format="%.1fx",
                  help="EV / EBITDA multiple at exit.\n\n"
                       "**Recommended:** S&P 500 long-run average: **~12x**. Mature large-caps: **8-12x**. "
                       "Growth companies: **15-25x**. Premium software / SaaS: **25-35x**. "
                       "Capital-intensive / cyclical (auto, airlines): **5-8x**.")

    st.markdown("---")

    # ─── WACC: direct or CAPM build-up ───────────────────────────────────────
    st.subheader("💸 Discount Rate (WACC)")
    with st.expander("ℹ️ Typical WACC ranges by risk", expanded=False):
        st.markdown(
            "- **Low risk** (large utilities, consumer staples): **5–7%**\n"
            "- **Medium risk** (large-cap consumer / industrial / healthcare): **7–10%**\n"
            "- **High risk** (tech / growth / small-cap): **10–14%**\n"
            "- **Very high risk** (early-stage / biotech / emerging markets): **14%+**\n\n"
            "*Higher beta = higher WACC. Higher debt-to-equity = lower WACC (debt is cheaper than equity).*"
        )
    st.radio("WACC mode", options=["Direct", "CAPM build-up"],
             key="wacc_mode", horizontal=True,
             help="Direct = enter WACC directly. CAPM build-up = compute WACC from risk-free rate, beta, equity risk premium, cost of debt, and capital structure (textbook formula).")

    if st.session_state.wacc_mode == "Direct":
        st.slider("WACC (%)", min_value=1.0, max_value=30.0, step=0.25, key="wacc",
                  format="%.2f%%",
                  help="Weighted Average Cost of Capital.\n\n"
                       "**Recommended:** **8-9%** for blue-chip large-caps (AAPL, MSFT, JNJ). "
                       "**9-11%** for mid-cap / cyclical (industrials, consumer discretionary). "
                       "**11-14%** for growth tech, smaller / international. **14%+** for biotech, micro-cap, "
                       "high-leverage, or emerging-market firms. If unsure, use **CAPM build-up** mode below.")
    else:
        st.slider("Risk-free Rate Rf (%)", min_value=0.0, max_value=10.0, step=0.05,
                  key="rf", format="%.2f%%",
                  help="10-year US Treasury yield.\n\n"
                       "**Recommended:** **~4.2%** (current 10-yr UST as of 2026). Use the actual current yield "
                       "from cnbc.com/quotes/US10Y or finance.yahoo.com (^TNX). For non-US firms, use the local "
                       "10-year sovereign yield + a country-risk premium.")
        st.slider("Equity Risk Premium ERP (%)", min_value=2.0, max_value=12.0, step=0.1,
                  key="erp", format="%.2f%%",
                  help="Excess return required over the risk-free rate for holding stocks.\n\n"
                       "**Recommended:** **5.5%** as a default (Damodaran current implied ERP for the S&P 500). "
                       "Historical realised: **~6%** (1928-2024). Forward-looking estimates from Wall Street: "
                       "**4.5-6.5%**. For emerging markets, add a country-risk premium of 1-5%.")
        st.slider("Beta (β)", min_value=0.0, max_value=3.0, step=0.05, key="beta",
                  help="Stock's sensitivity to market moves. 1.0 = market average.\n\n"
                       "**Recommended:** look up on finance.yahoo.com (Statistics tab) or stockanalysis.com. "
                       "Defensive / staples / utilities: **0.5-0.9**. Broad market average: **1.0**. "
                       "Tech / cyclicals: **1.1-1.5**. High-beta names (NVDA, TSLA, semis, biotech): **1.5-2.5**.")
        st.slider("Pre-tax Cost of Debt Rd (%)", min_value=0.0, max_value=15.0, step=0.1,
                  key="rd_pretax", format="%.2f%%",
                  help="Yield on the company's debt (pre-tax).\n\n"
                       "**Recommended:** AAA / large-cap blue chip: **4-5%**. Investment-grade (BBB to AA): **5-6.5%**. "
                       "BB / weaker IG: **6.5-8%**. High-yield (B and below): **8-12%**. "
                       "Approximate as Risk-free + credit spread (look up on FRED's BAML indices).")
        st.slider("Target Debt Weight (D/V) (%)", min_value=0.0, max_value=80.0, step=1.0,
                  key="weight_debt", format="%.0f%%",
                  help="Debt as % of total capital. Equity weight = 100% − Debt weight.\n\n"
                       "**Recommended:** tech / asset-light: **5-15%**. Healthcare / consumer staples: **20-30%**. "
                       "Industrials / energy: **30-40%**. Utilities / REITs / banks: **40-60%**. "
                       "Use **market-value** weights (book values are often misleading). "
                       "Approximate: market cap ÷ (market cap + total debt).")

    st.markdown("---")

    # ─── Advanced settings ───────────────────────────────────────────────────
    with st.expander("🔧 Advanced settings"):
        st.checkbox("Mid-year discounting convention", key="mid_year",
                    help="Standard practice in banking: assumes cash flows arrive mid-year, "
                         "discounted with exponent (t − 0.5) instead of t. Yields ~half-year-of-WACC higher valuation.")
        st.slider("Verdict threshold — 'Fairly Valued' band (±%)", min_value=2.0, max_value=25.0,
                  step=1.0, key="fair_band", format="%.0f%%",
                  help="If intrinsic value is within ±this band of market price, verdict says 'Fairly Valued'. "
                       "Outside the band: 'Undervalued' or 'Overvalued'.\n\n"
                       "**Recommended:** **10%** as a balanced default. Tighter (5%) is too strict given DCF "
                       "input uncertainty. Wider (20%+) makes most stocks look 'fairly valued' and weakens the signal.")

    # Capital Structure (Total Debt, Cash & Equivalents, Shares Outstanding)
    # and Market Reference (Current Stock Price) widgets intentionally hidden.
    # debt/cash/shares are populated by EDGAR autofill on ticker change;
    # current_price keeps its DEFAULTS fallback. All four flow into the DCF
    # math from session_state so the customer sees the valuation, not the
    # raw inputs.


# =============================================================================
# Pull values out of session state for clean local references
# =============================================================================
ticker        = st.session_state.ticker
company_name  = st.session_state.company_name
revenue       = float(st.session_state.revenue)
growth        = float(st.session_state.growth) / 100
margin        = float(st.session_state.margin) / 100
tax_rate      = float(st.session_state.tax_rate) / 100
reinvest_rate = float(st.session_state.reinvest_rate) / 100
term_g        = float(st.session_state.terminal_growth) / 100
years         = int(st.session_state.years)
debt          = float(st.session_state.debt)
cash          = float(st.session_state.cash)
shares        = max(float(st.session_state.shares), 0.1)  # never zero (divide guard)
price         = max(float(st.session_state.current_price), 0.01)

# Resolve WACC (direct entry OR CAPM build-up) BEFORE validation
cost_equity = after_tax_rd = None
if st.session_state.wacc_mode == "CAPM build-up":
    wacc, cost_equity, after_tax_rd = None, None, None  # filled below
else:
    wacc = float(st.session_state.wacc) / 100


# =============================================================================
# DCF Calculation — accepts per-year growth/margin paths for multi-stage models
# =============================================================================
def build_growth_path(simple_g, n, mode, hg_rate, hg_years, fade_years, term_g):
    """Generate a per-year growth-rate list.
    Modes:
      "Constant": flat g for all forecast years.
      "3-stage": high-growth period at hg_rate, then linear fade over fade_years
                 down to term_g, then term_g for any remaining years.
    """
    if mode != "3-stage (high → fade → terminal)":
        return [simple_g] * n
    path = []
    hg_yrs    = max(0, int(hg_years))
    fade_yrs  = max(1, int(fade_years))
    for t in range(1, n + 1):
        if t <= hg_yrs:
            path.append(hg_rate)
        elif t <= hg_yrs + fade_yrs:
            # Linear interpolation from hg_rate (start of fade) to term_g (end)
            step = (t - hg_yrs) / fade_yrs
            path.append(hg_rate + (term_g - hg_rate) * step)
        else:
            path.append(term_g)
    return path


def build_margin_path(simple_m, n, mode, terminal_m):
    """Generate a per-year EBIT margin list.
    Modes:
      "Constant": flat margin for all years.
      "Linear expansion": linear interpolation from simple_m (Year 1) to terminal_m (Year N).
    """
    if mode != "Linear expansion" or n < 2:
        return [simple_m] * n
    return [simple_m + (terminal_m - simple_m) * (t - 1) / (n - 1) for t in range(1, n + 1)]


def compute_capm_wacc(rf, beta, erp, rd_pretax, w_debt, tx):
    """Compute WACC from CAPM components.
    cost_equity   = rf + beta × ERP
    after_tax_rd  = rd_pretax × (1 − tax)
    WACC          = w_equity × cost_equity + w_debt × after_tax_rd
    Returns (wacc, cost_equity, after_tax_rd) — all decimals.
    """
    cost_equity  = rf + beta * erp
    after_tax_rd = rd_pretax * (1 - tx)
    w_equity     = 1 - w_debt
    wacc         = w_equity * cost_equity + w_debt * after_tax_rd
    return wacc, cost_equity, after_tax_rd


def run_dcf(rev0, growth_path, margin_path, tx, ri, w, tg, n,
            *, mid_year=False, tv_method="Gordon Growth", exit_multiple=10.0,
            nwc_pct=0.0, prev_revenue=None):
    """Run a DCF projection with per-year growth and margin paths.

    growth_path[t] applies to Year t+1 (compounded onto prior year's revenue).
    margin_path[t] is the EBIT margin for Year t+1.
    nwc_pct is change in net working capital as % of revenue change (deducted from FCF).
    mid_year=True uses (t-0.5) exponent for discounting (assumes cash arrives mid-year).
    tv_method='Exit Multiple' computes terminal value as exit_multiple × EBITDA_N
        (we approximate EBITDA ≈ EBIT × 1.15 since D&A isn't separately modelled here).
    """
    revenues, ebits, nopats, reinvs, nwcs, fcfs, dfs, pvs = [], [], [], [], [], [], [], []
    rev = rev0
    prev_rev = prev_revenue if prev_revenue is not None else rev0
    for t in range(n):
        g_t = growth_path[t]
        m_t = margin_path[t]
        new_rev = rev * (1 + g_t)
        d_rev   = new_rev - prev_rev
        rev     = new_rev
        ebit    = rev * m_t
        nopat   = ebit * (1 - tx)
        reinv   = nopat * ri
        nwc     = max(0, d_rev * nwc_pct)  # only deduct if revenue grew
        fcf     = nopat - reinv - nwc
        # Discounting: end-year (t) or mid-year (t - 0.5)
        exp = (t + 0.5) if mid_year else (t + 1)
        df = 1 / (1 + w) ** exp
        pv = fcf * df
        revenues.append(rev); ebits.append(ebit); nopats.append(nopat)
        reinvs.append(reinv); nwcs.append(nwc); fcfs.append(fcf); dfs.append(df); pvs.append(pv)
        prev_rev = rev

    sum_pv_fcf = sum(pvs)
    last_fcf  = fcfs[-1] if fcfs else 0
    last_ebit = ebits[-1] if ebits else 0
    if tv_method == "Exit Multiple":
        # EBITDA ≈ EBIT × 1.15 as a rough proxy (no D&A line modelled here)
        terminal_value = exit_multiple * (last_ebit * 1.15)
        terminal_fcf   = last_fcf * (1 + tg)  # only used for display
    else:
        terminal_fcf   = last_fcf * (1 + tg)
        terminal_value = terminal_fcf / (w - tg) if w > tg else float("nan")
    # Discount terminal value back: end-year uses (1+w)^n, mid-year uses (1+w)^(n-0.5)
    tv_exp = (n - 0.5) if mid_year else n
    pv_terminal = terminal_value / (1 + w) ** tv_exp if n > 0 else 0
    return {
        "revenues": revenues, "ebits": ebits, "nopats": nopats, "reinvs": reinvs,
        "nwcs": nwcs, "fcfs": fcfs, "dfs": dfs, "pvs": pvs,
        "growth_path": growth_path, "margin_path": margin_path,
        "sum_pv_fcf": sum_pv_fcf,
        "terminal_fcf": terminal_fcf,
        "terminal_value": terminal_value,
        "pv_terminal": pv_terminal,
        "enterprise_value": sum_pv_fcf + pv_terminal,
        "tv_method": tv_method,
    }


# Build growth + margin paths from the user's choice of mode
growth_path = build_growth_path(
    growth, years, st.session_state.growth_mode,
    st.session_state.high_growth / 100,
    st.session_state.high_years,
    st.session_state.fade_years,
    term_g,
)
margin_path = build_margin_path(
    margin, years, st.session_state.margin_mode,
    st.session_state.margin_terminal / 100,
)

# Compute WACC — either from CAPM components or use direct entry
if st.session_state.wacc_mode == "CAPM build-up":
    wacc, cost_equity, after_tax_rd = compute_capm_wacc(
        st.session_state.rf / 100,
        st.session_state.beta,
        st.session_state.erp / 100,
        st.session_state.rd_pretax / 100,
        st.session_state.weight_debt / 100,
        tax_rate,
    )

# Now validate (WACC has been resolved either way)
input_errors = []
if revenue <= 0:
    input_errors.append("Revenue must be > 0.")
if margin <= 0:
    input_errors.append("EBIT Margin must be > 0% (the firm needs operating profit for DCF to work).")
if wacc <= term_g:
    input_errors.append(
        f"WACC ({wacc*100:.2f}%) must be greater than Terminal Growth ({term_g*100:.2f}%) — "
        "otherwise the Gordon Growth formula explodes to infinity."
    )
if shares <= 0:
    input_errors.append("Shares Outstanding must be > 0.")
if input_errors:
    st.error("⚠️ **Cannot run DCF — please fix these inputs first:**\n\n" +
             "\n".join(f"- {e}" for e in input_errors))
    st.stop()

dcf = run_dcf(
    revenue, growth_path, margin_path, tax_rate, reinvest_rate, wacc, term_g, years,
    mid_year=st.session_state.mid_year,
    tv_method=st.session_state.tv_method,
    exit_multiple=st.session_state.exit_multiple,
    nwc_pct=st.session_state.nwc_pct / 100,
)

enterprise_value = dcf["enterprise_value"]
equity_value     = enterprise_value - debt + cash
intrinsic_per    = equity_value / shares
upside           = (intrinsic_per - price) / price if price > 0 else 0
mos              = (intrinsic_per - price) / intrinsic_per if intrinsic_per > 0 else 0


# =============================================================================
# Page header — company + headline metrics
# =============================================================================
# If the loaded company name doesn't match the current ticker (user typed a new
# ticker but autofill hasn't refreshed yet), show ticker only — never display a
# stale company name that doesn't match the ticker.
_loaded = (st.session_state.get("loaded_ticker") or "").strip().upper()
_cur_t = (ticker or "").strip().upper()
if _loaded == _cur_t and company_name:
    st.markdown(f"## 🏢 {company_name} ({ticker})")
else:
    st.markdown(f"## 🏢 {ticker}")

m1, m2, m3, m4 = st.columns(4)
m1.metric("DCF Intrinsic Value", f"${intrinsic_per:,.2f}")
m2.metric("Current Market Price", f"${price:,.2f}")
m3.metric("Upside / (Downside)", f"{upside:+.1%}",
          delta=f"${intrinsic_per - price:,.2f}",
          delta_color="normal" if abs(upside) < 0.05 else ("inverse" if upside < 0 else "normal"))
m4.metric("Margin of Safety", f"{mos:+.1%}")

# Verdict box (threshold customisable in Advanced settings)
fair_band = st.session_state.fair_band / 100
if abs(upside) < fair_band:
    st.info(f"⚖️  **Fairly Valued** — DCF intrinsic value is within ±{fair_band*100:.0f}% of market price.")
elif upside > 0:
    st.success(f"📈 **Potentially Undervalued** — DCF estimates the stock is worth {upside:.1%} more than its current price.")
else:
    st.warning(f"📉 **Potentially Overvalued** — Market is paying {-upside:.1%} above the DCF estimate.")

# Hint when DCF is meaningfully below market — common for premium-priced stocks
if upside < -0.30:
    st.caption(
        "💡 *DCF often produces lower values than market price for high-multiple stocks. "
        "The market may be pricing in faster growth, margin expansion, or longer-than-modeled growth. "
        "Try increasing **growth rate**, increasing **EBIT margin**, lowering **WACC**, or extending the **forecast horizon** "
        "to model a more bullish view. Use the sensitivity tables to see what assumptions justify the market price.*"
    )

# Implied multiples — what valuation multiples your DCF assumptions produce
# Useful sanity check vs how the market is currently pricing the stock
_market_cap = price * shares
_implied_pe = intrinsic_per / (dcf["nopats"][0] / shares) if dcf["nopats"] and dcf["nopats"][0] > 0 else None
_implied_ev_ebit = enterprise_value / dcf["ebits"][0] if dcf["ebits"] and dcf["ebits"][0] > 0 else None
_market_pe = price / (dcf["nopats"][0] / shares) if dcf["nopats"] and dcf["nopats"][0] > 0 else None
with st.expander("📐 Implied valuation multiples (sanity check)", expanded=False):
    im1, im2, im3 = st.columns(3)
    if _implied_pe and _market_pe:
        im1.metric("Your DCF implies P/E of", f"{_implied_pe:.1f}x",
                   delta=f"vs market {_market_pe:.1f}x", delta_color="off")
    if _implied_ev_ebit:
        im2.metric("Your DCF implies EV/EBIT of", f"{_implied_ev_ebit:.1f}x")
    im3.metric("Implied Market Cap", f"${(intrinsic_per * shares)/1000:,.1f}B",
               delta=f"vs current ${(_market_cap)/1000:,.1f}B")
    st.caption(
        "*A reasonable DCF for a mature firm typically implies a P/E of 12-25x. Above 30x usually means "
        "your growth/margin assumptions are aggressive; below 10x suggests they're conservative.*"
    )


# =============================================================================
# CAPM WACC build-up panel (only when CAPM mode is active)
# =============================================================================
if st.session_state.wacc_mode == "CAPM build-up":
    with st.expander(f"💸 WACC build-up (CAPM) → **{wacc*100:.2f}%**", expanded=False):
        st.latex(r"\text{Cost of Equity} = R_f + \beta \times \text{ERP}")
        st.latex(r"\text{WACC} = \frac{E}{V} \times \text{Cost of Equity} + \frac{D}{V} \times R_d \times (1 - \text{tax})")
        c1, c2, c3 = st.columns(3)
        c1.metric("Risk-free Rate (Rf)", f"{st.session_state.rf:.2f}%")
        c1.metric("Beta (β)", f"{st.session_state.beta:.2f}")
        c2.metric("Equity Risk Premium", f"{st.session_state.erp:.2f}%")
        c2.metric("Cost of Equity (Re)", f"{cost_equity*100:.2f}%")
        c3.metric("Pre-tax Cost of Debt", f"{st.session_state.rd_pretax:.2f}%")
        c3.metric("After-tax Cost of Debt", f"{after_tax_rd*100:.2f}%")
        st.markdown(
            f"**Capital structure:** {(100-st.session_state.weight_debt):.0f}% Equity / "
            f"{st.session_state.weight_debt:.0f}% Debt  →  "
            f"**WACC = {wacc*100:.2f}%**"
        )


# =============================================================================
# Multi-stage growth / margin path display (only when non-constant)
# =============================================================================
if st.session_state.growth_mode != "Constant" or st.session_state.margin_mode != "Constant":
    with st.expander("📐 Growth & Margin Path (per year)", expanded=False):
        path_df = pd.DataFrame({
            "Year":             [f"Year {t+1}" for t in range(years)],
            "Growth Rate":      [f"{g*100:.2f}%" for g in growth_path],
            "EBIT Margin":      [f"{m*100:.2f}%" for m in margin_path],
        })
        st.dataframe(path_df, hide_index=True, use_container_width=True)
        if st.session_state.growth_mode != "Constant":
            st.caption(
                f"3-stage growth: {st.session_state.high_years} years at "
                f"{st.session_state.high_growth:.1f}%, then {st.session_state.fade_years}-year "
                f"linear fade to {st.session_state.terminal_growth:.2f}% terminal."
            )
        if st.session_state.margin_mode != "Constant":
            st.caption(
                f"Linear margin expansion: {st.session_state.margin:.1f}% → "
                f"{st.session_state.margin_terminal:.1f}% over {years} years."
            )


# =============================================================================
# EDGAR context box (if data was loaded)
# =============================================================================
if st.session_state.edgar_data:
    e = st.session_state.edgar_data
    with st.expander("📚 SEC EDGAR Reference Data (most recent 10-K)", expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**5-Year Revenue History (from 10-K filings):**")
            if e.get("rev_history"):
                hist_df = pd.DataFrame(e["rev_history"], columns=["FY ending", "Revenue ($B)"])
                st.dataframe(hist_df, hide_index=True, use_container_width=True)
        with c2:
            st.markdown("**Computed metrics:**")
            def _fmt_pct(v): return f"{v:.1f}%" if isinstance(v, (int, float)) else "N/A"
            def _fmt_b(v):   return f"${v/1000:.1f}B" if isinstance(v, (int, float)) else "N/A"
            ocf_v   = e.get("operating_cf")
            capex_v = e.get("capex")
            fcf_v   = (ocf_v - capex_v) if (isinstance(ocf_v, (int, float)) and isinstance(capex_v, (int, float))) else None
            st.write(f"- 3-yr Revenue CAGR: **{_fmt_pct(e.get('revenue_cagr'))}**")
            st.write(f"- Operating Cash Flow: **{_fmt_b(ocf_v)}**")
            st.write(f"- CapEx: **{_fmt_b(capex_v)}**")
            st.write(f"- Implied FCF: **{_fmt_b(fcf_v)}**")
        st.caption("Source: data.sec.gov XBRL Company Facts API · Used as starting point — every input above is editable.")


st.markdown("---")


# =============================================================================
# Step-by-step walkthrough
# =============================================================================
st.markdown("## 🧮 Step-by-Step Valuation")
st.markdown(
    "*The DCF method values a company by **forecasting the cash it will generate**, "
    "then **discounting those cash flows back to today** at the firm's cost of capital. "
    "A dollar tomorrow is worth less than a dollar today — this app shows how much less.*"
)

# ─── Step 1: Forecast ──────────────────────────────────────────────────────
st.markdown(f"### Step 1 — Forecast Free Cash Flows (Years 1 to {years})")
st.markdown(
    "We project revenue forward at the assumed growth rate, apply the EBIT margin to get "
    "operating profit, take out taxes to get NOPAT, then deduct what the firm must reinvest "
    "to support that growth. **What's left is free cash flow** — the cash that could be returned to investors."
)
st.markdown("**Formula:**")
st.latex(r"\text{Revenue}_t = \text{Revenue}_{t-1} \times (1 + g)")
st.latex(r"\text{EBIT}_t = \text{Revenue}_t \times \text{margin}")
st.latex(r"\text{NOPAT}_t = \text{EBIT}_t \times (1 - \text{tax rate})")
st.latex(r"\text{FCF}_t = \text{NOPAT}_t \times (1 - \text{reinvestment rate})")

proj_df = pd.DataFrame({
    "Year":           list(range(1, years + 1)),
    "Growth":         [f"{g*100:.2f}%" for g in dcf["growth_path"]],
    "Margin":         [f"{m*100:.2f}%" for m in dcf["margin_path"]],
    "Revenue ($M)":   [f"{x:,.0f}" for x in dcf["revenues"]],
    "EBIT ($M)":      [f"{x:,.0f}" for x in dcf["ebits"]],
    "NOPAT ($M)":     [f"{x:,.0f}" for x in dcf["nopats"]],
    "Reinvest ($M)":  [f"{x:,.0f}" for x in dcf["reinvs"]],
    "Δ NWC ($M)":     [f"{x:,.0f}" for x in dcf["nwcs"]],
    "FCF ($M)":       [f"{x:,.0f}" for x in dcf["fcfs"]],
})
st.dataframe(proj_df, hide_index=True, use_container_width=True)

# Chart
fig = go.Figure()
fig.add_trace(go.Bar(x=list(range(1, years + 1)), y=dcf["fcfs"], name="FCF",
                      marker_color="#1f4e79"))
fig.update_layout(
    title="Forecasted Free Cash Flows",
    xaxis_title="Year", yaxis_title="FCF ($M)",
    template="plotly_white", height=350, margin=dict(t=50, b=40),
)
st.plotly_chart(fig, use_container_width=True)


# ─── Step 2: Discount ──────────────────────────────────────────────────────
st.markdown(f"### Step 2 — Discount FCFs to Present Value (WACC = {wacc*100:.2f}%)")
st.markdown(
    "Each future cash flow is worth less than a dollar received today. We discount each year's "
    f"FCF by `1 / (1 + WACC)^t` — at a {wacc*100:.2f}% cost of capital, $1 received in 5 years is "
    f"worth about **${(1/(1+wacc)**5):.2f}** today."
)
st.markdown("**Formula:**")
st.latex(r"\text{PV of FCF}_t = \frac{\text{FCF}_t}{(1 + \text{WACC})^t}")

disc_df = pd.DataFrame({
    "Year":              list(range(1, years + 1)),
    "FCF ($M)":          [f"{x:,.0f}" for x in dcf["fcfs"]],
    "Discount Factor":   [f"{x:.4f}" for x in dcf["dfs"]],
    "PV of FCF ($M)":    [f"{x:,.0f}" for x in dcf["pvs"]],
})
st.dataframe(disc_df, hide_index=True, use_container_width=True)
st.markdown(f"**Sum of PV of FCFs = ${dcf['sum_pv_fcf']:,.0f}M**")


# ─── Step 3: Terminal Value ────────────────────────────────────────────────
st.markdown("### Step 3 — Terminal Value")
st.markdown(
    "We can't forecast forever. The **terminal value** captures everything beyond the explicit "
    "forecast horizon — assuming the firm grows steadily at a perpetual rate (Gordon Growth) "
    "or could be sold at a market multiple (Exit Multiple)."
)
st.markdown("**Formula:**")
st.latex(r"\text{Terminal Value} = \frac{\text{FCF}_{N} \times (1 + g_{\text{terminal}})}{\text{WACC} - g_{\text{terminal}}}")
st.latex(r"\text{PV of Terminal Value} = \frac{\text{Terminal Value}}{(1 + \text{WACC})^N}")

t1, t2, t3 = st.columns(3)
t1.metric("Terminal FCF (Yr N+1)", f"${dcf['terminal_fcf']:,.0f}M")
t2.metric("Terminal Value", f"${dcf['terminal_value']:,.0f}M")
t3.metric("PV of Terminal Value", f"${dcf['pv_terminal']:,.0f}M")

# Composition of enterprise value: explicit FCF vs terminal — a key sanity check
_proj_pct = dcf["sum_pv_fcf"] / dcf["enterprise_value"] if dcf["enterprise_value"] else 0
_term_pct = 1 - _proj_pct
st.markdown(
    f"📌 **Composition of Enterprise Value:** "
    f"`{_proj_pct:.0%}` from explicit-period FCFs · "
    f"`{_term_pct:.0%}` from terminal value. "
    f"{'Terminal value over 70-80% is normal for short forecast horizons.' if _term_pct < 0.85 else '⚠️ A very high terminal share (>85%) means valuation is highly sensitive to perpetual-growth assumptions.'}"
)
if dcf['pv_terminal'] / dcf['enterprise_value'] > 0.85:
    st.warning(
        f"Consider extending the forecast horizon or revisiting the terminal growth rate."
    )


# ─── Step 4: Enterprise Value ──────────────────────────────────────────────
st.markdown("### Step 4 — Enterprise Value")
st.markdown(
    "Adding it all together: the present value of explicit-period free cash flows + the present "
    "value of everything that comes after. This is the value of the **operating business** to all "
    "capital providers (equity + debt holders)."
)
st.markdown("**Formula:**")
st.latex(r"\text{EV} = \sum \text{PV of FCFs} + \text{PV of Terminal Value}")

e1, e2, e3 = st.columns(3)
e1.metric("Sum of PV of FCFs", f"${dcf['sum_pv_fcf']:,.0f}M")
e2.metric("PV of Terminal Value", f"${dcf['pv_terminal']:,.0f}M")
e3.metric("Enterprise Value", f"${enterprise_value:,.0f}M")


# ─── Step 5: Equity Value ──────────────────────────────────────────────────
st.markdown("### Step 5 — Bridge to Equity Value")
st.markdown(
    "Equity holders get what's left **after debt is paid off** but they also benefit from any "
    "excess cash on the balance sheet. Subtract debt, add cash, divide by diluted shares — "
    "and we have intrinsic value per share."
)
st.markdown("**Formula:**")
st.latex(r"\text{Equity Value} = \text{EV} - \text{Debt} + \text{Cash}")
st.latex(r"\text{Value per Share} = \frac{\text{Equity Value}}{\text{Shares Outstanding}}")

bridge_df = pd.DataFrame({
    "Item":         ["Enterprise Value", "(−) Total Debt", "(+) Cash & Equivalents", "= Equity Value", "÷ Shares Outstanding (M)", "= Value per Share"],
    "Value ($M / $)": [
        f"${enterprise_value:,.0f}M",
        f"−${debt:,.0f}M",
        f"+${cash:,.0f}M",
        f"${equity_value:,.0f}M",
        f"{shares:,.0f}M shares",
        f"${intrinsic_per:,.2f}",
    ],
})
st.dataframe(bridge_df, hide_index=True, use_container_width=True)


# ─── Step 6: Compare to market ─────────────────────────────────────────────
st.markdown("### Step 6 — Compare to Market Price")

cmp_fig = go.Figure()
cmp_fig.add_trace(go.Bar(x=["Market Price", "DCF Intrinsic Value"],
                          y=[price, intrinsic_per],
                          marker_color=["#999999", "#1f4e79"],
                          text=[f"${price:.2f}", f"${intrinsic_per:.2f}"],
                          textposition="outside"))
cmp_fig.update_layout(
    title=f"{ticker} — Market vs. Intrinsic", yaxis_title="Price (USD)",
    template="plotly_white", height=400, showlegend=False, margin=dict(t=50, b=40),
)
st.plotly_chart(cmp_fig, use_container_width=True)


st.markdown("---")


# =============================================================================
# Sensitivity analysis
# =============================================================================
st.markdown("## 📊 Sensitivity Analysis")
st.caption("Intrinsic value per share under different assumption combinations. "
           "Color scale: 🔴 lowest value → 🟡 mid → 🟢 highest value within each table. "
           "**Bordered cell** = your current base case.")

# Smooth red→yellow→green colormap that works on both light and dark themes.
# Uses HSL: red (0°) → yellow (60°) → green (120°), with low saturation to keep
# text readable. Returns CSS background strings for pandas Styler.applymap.
def _gradient_style(val, vmin, vmax):
    if pd.isna(val) or vmax == vmin:
        return ""
    ratio = max(0, min(1, (val - vmin) / (vmax - vmin)))
    hue = int(120 * ratio)              # 0 = red, 60 = yellow, 120 = green
    return f"background-color: hsl({hue}, 65%, 78%); color: #111;"

def style_table_gradient(df, base_row=None, base_col=None):
    """Apply a per-table red→green gradient based on numeric cell values.
    Min value in the table → red, max → green, linear interpolation between.
    If base_row + base_col are given, that cell gets a thick navy border to
    mark it as the user's current assumptions (base case).
    """
    flat = df.stack().dropna()
    if flat.empty:
        return df.style.format("${:,.2f}")
    vmin, vmax = float(flat.min()), float(flat.max())

    def _cell_style(v):
        return _gradient_style(v, vmin, vmax) if pd.notna(v) else ""

    styler = df.style.format("${:,.2f}").map(_cell_style)
    # Thick border on the base-case cell
    if base_row is not None and base_col is not None and base_row in df.index and base_col in df.columns:
        styler = styler.set_properties(
            subset=pd.IndexSlice[[base_row], [base_col]],
            **{"border": "3px solid #1F4E79", "font-weight": "bold"},
        )
    return styler

# Table 1: WACC × Terminal Growth
# Run DCF with given overrides; reuses the live growth/margin paths and all advanced settings
def _sens_dcf(rev=None, g=None, m=None, w=None, tg=None):
    rev_use = revenue if rev is None else rev
    g_use   = growth if g is None else g
    m_use   = margin if m is None else m
    w_use   = wacc if w is None else w
    tg_use  = term_g if tg is None else tg
    # Rebuild paths with the overridden simple growth/margin
    gp = build_growth_path(g_use, years, st.session_state.growth_mode,
                           st.session_state.high_growth/100,
                           st.session_state.high_years, st.session_state.fade_years, tg_use)
    mp = build_margin_path(m_use, years, st.session_state.margin_mode,
                           st.session_state.margin_terminal/100)
    return run_dcf(rev_use, gp, mp, tax_rate, reinvest_rate, w_use, tg_use, years,
                   mid_year=st.session_state.mid_year,
                   tv_method=st.session_state.tv_method,
                   exit_multiple=st.session_state.exit_multiple,
                   nwc_pct=st.session_state.nwc_pct/100)

st.markdown("### Table 1 — Intrinsic Value: WACC vs. Terminal Growth")
wacc_range = [round(wacc + dx, 4) for dx in [-0.02, -0.01, 0, 0.01, 0.02]]
tg_range   = [round(term_g + dx, 4) for dx in [-0.01, -0.005, 0, 0.005, 0.01]]
tbl1 = []
for w in wacc_range:
    if w <= 0: continue
    row = []
    for tg in tg_range:
        if tg < 0 or w <= tg:
            row.append(np.nan); continue
        d = _sens_dcf(w=w, tg=tg)
        eq = d["enterprise_value"] - debt + cash
        row.append(eq / shares)
    tbl1.append(row)
df1 = pd.DataFrame(
    tbl1,
    index=[f"WACC={w*100:.2f}%" for w in wacc_range if w > 0],
    columns=[f"g={tg*100:.2f}%" for tg in tg_range],
)
_base_w_label  = f"WACC={wacc*100:.2f}%"
_base_tg_label = f"g={term_g*100:.2f}%"
st.dataframe(style_table_gradient(df1, base_row=_base_w_label, base_col=_base_tg_label),
             use_container_width=True)

# Table 2: Growth × Margin
st.markdown("### Table 2 — Intrinsic Value: Growth Rate vs. EBIT Margin")
g_range = [round(growth + dx, 4) for dx in [-0.04, -0.02, 0, 0.02, 0.04]]
m_range = [round(margin + dx, 4) for dx in [-0.05, -0.025, 0, 0.025, 0.05]]
tbl2 = []
for g in g_range:
    row = []
    for m in m_range:
        if m <= 0: row.append(np.nan); continue
        d = _sens_dcf(g=g, m=m)
        eq = d["enterprise_value"] - debt + cash
        row.append(eq / shares)
    tbl2.append(row)
df2 = pd.DataFrame(
    tbl2,
    index=[f"g={g*100:.2f}%" for g in g_range],
    columns=[f"margin={m*100:.2f}%" for m in m_range],
)
_base_g_label  = f"g={growth*100:.2f}%"
_base_m_label  = f"margin={margin*100:.2f}%"
st.dataframe(style_table_gradient(df2, base_row=_base_g_label, base_col=_base_m_label),
             use_container_width=True)


st.markdown("---")


# =============================================================================
# Excel export — download a workbook that REPLICATES the DCF using formulas
# =============================================================================
st.markdown("## 📥 Export to Excel")
st.caption(
    "Download a fully editable Excel workbook with live formulas — every projection cell "
    "references the inputs sheet, so you can modify any assumption and the entire model recomputes."
)

def build_excel_workbook():
    """Build a clean Excel workbook that demonstrates the DCF math step by step.

    Two sheets — Inputs and DCF Model — designed so anyone opening the file
    can trace every formula and verify the valuation by hand. Live formulas
    throughout: change any input → entire model recomputes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY       = "1F4E79"
    LIGHT_BLUE = "D9E2F3"
    SOFT_GREEN = "C5E0B4"
    INPUT_BLUE = "0070C0"

    f_title  = Font(name="Calibri", size=18, bold=True, color=NAVY)
    f_sub    = Font(name="Calibri", size=11, italic=True, color="595959")
    f_hdr    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    f_bold   = Font(name="Calibri", size=10, bold=True)
    f_input  = Font(name="Calibri", size=10, color=INPUT_BLUE)
    f_summary = Font(name="Calibri", size=12, bold=True, color=NAVY)

    fill_hdr = PatternFill("solid", fgColor=NAVY)
    fill_sec = PatternFill("solid", fgColor=LIGHT_BLUE)
    fill_sum = PatternFill("solid", fgColor=SOFT_GREEN)

    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="404040")
    box   = Border(left=thin, right=thin, top=thin, bottom=thin)
    high  = Border(left=thin, right=thin, top=thick, bottom=thick)

    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    indent = Alignment(horizontal="left",   vertical="center", indent=1)

    fmt_pct    = "0.00%"
    fmt_dollar = "$#,##0;($#,##0)"
    fmt_dol2   = "$#,##0.00;($#,##0.00)"
    fmt_int    = "#,##0;(#,##0)"
    fmt_factor = "0.0000"

    wb = Workbook()

    # =====================================================================
    # SHEET 1 — Inputs (all assumptions in one place, every cell editable)
    # =====================================================================
    ws = wb.active
    ws.title = "Inputs"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "DCF Model — Inputs"; ws["A1"].font = f_title
    ws["A2"] = f"{company_name} ({ticker})  |  Edit blue cells to change assumptions"
    ws["A2"].font = f_sub
    ws.merge_cells("A1:C1"); ws.merge_cells("A2:C2")

    # Header row
    ws["A4"] = "Input"; ws["B4"] = "Value"; ws["C4"] = "Unit"
    for col in ["A4", "B4", "C4"]:
        ws[col].font = f_hdr; ws[col].fill = fill_hdr
        ws[col].alignment = center; ws[col].border = box
    ws.row_dimensions[4].height = 22

    # All inputs the model needs to MATCH the website's output exactly.
    # When the website has advanced features active (3-stage growth, margin
    # expansion, mid-year, exit multiple, NWC drag), we write those values
    # here so the Excel produces the same intrinsic value.
    inputs = [
        ("Current Revenue",          revenue,                                                   "$M",  fmt_dollar),
        ("Tax Rate",                 tax_rate,                                                  "%",   fmt_pct),
        ("Reinvestment Rate",        reinvest_rate,                                             "%",   fmt_pct),
        ("ΔNWC (% of Δrevenue)",     st.session_state.nwc_pct/100,                              "%",   fmt_pct),
        ("WACC (Discount Rate)",     wacc,                                                      "%",   fmt_pct),
        ("Terminal Growth Rate",     term_g,                                                    "%",   fmt_pct),
        ("Forecast Years",           years,                                                     "yrs", "0"),
        ("Mid-year discounting (1/0)", 1 if st.session_state.mid_year else 0,                   "",    "0"),
        ("Exit Multiple (1/0)",      1 if st.session_state.tv_method=="Exit Multiple" else 0,   "",    "0"),
        ("Exit EV/EBITDA Multiple",  st.session_state.exit_multiple,                            "x",   "0.0"),
        ("Total Debt",               debt,                                                      "$M",  fmt_dollar),
        ("Cash & Equivalents",       cash,                                                      "$M",  fmt_dollar),
        ("Shares Outstanding",       shares,                                                    "M",   fmt_int),
        ("Current Stock Price",      price,                                                     "$",   fmt_dol2),
    ]
    INPUT_ROW = {}
    cur = 5
    for label, val, unit, fmt in inputs:
        c1 = ws.cell(row=cur, column=1, value=label)
        c1.font = f_bold; c1.alignment = indent; c1.border = box
        c2 = ws.cell(row=cur, column=2, value=val)
        c2.font = f_input; c2.alignment = right; c2.border = box; c2.number_format = fmt
        c3 = ws.cell(row=cur, column=3, value=unit)
        c3.alignment = center; c3.border = box
        INPUT_ROW[label] = cur
        cur += 1

    # Per-year growth & margin path table — captures Constant / 3-stage / Linear
    # expansion modes. Whatever the website's active mode produced is written here;
    # user can override any single year.
    cur += 1
    ws.cell(row=cur, column=1, value="Per-Year Growth & Margin").font = f_bold
    cur += 1
    ws.cell(row=cur, column=1, value="Year").font = f_hdr
    ws.cell(row=cur, column=2, value="Growth Rate").font = f_hdr
    ws.cell(row=cur, column=3, value="EBIT Margin").font = f_hdr
    for c in range(1, 4):
        ws.cell(row=cur, column=c).fill = fill_hdr
        ws.cell(row=cur, column=c).alignment = center
        ws.cell(row=cur, column=c).border = box
    PATH_FIRST_ROW = cur + 1
    for t in range(years):
        r = PATH_FIRST_ROW + t
        c1 = ws.cell(row=r, column=1, value=f"Year {t+1}")
        c1.font = f_bold; c1.alignment = center; c1.border = box
        c2 = ws.cell(row=r, column=2, value=growth_path[t])
        c2.font = f_input; c2.alignment = right; c2.border = box; c2.number_format = fmt_pct
        c3 = ws.cell(row=r, column=3, value=margin_path[t])
        c3.font = f_input; c3.alignment = right; c3.border = box; c3.number_format = fmt_pct

    # Cell references
    REV_CELL    = f"Inputs!$B${INPUT_ROW['Current Revenue']}"
    TAX_CELL    = f"Inputs!$B${INPUT_ROW['Tax Rate']}"
    REINV_CELL  = f"Inputs!$B${INPUT_ROW['Reinvestment Rate']}"
    NWC_CELL    = f"Inputs!$B${INPUT_ROW['ΔNWC (% of Δrevenue)']}"
    WACC_CELL   = f"Inputs!$B${INPUT_ROW['WACC (Discount Rate)']}"
    TG_CELL     = f"Inputs!$B${INPUT_ROW['Terminal Growth Rate']}"
    YEARS_CELL  = f"Inputs!$B${INPUT_ROW['Forecast Years']}"
    MIDYR_CELL  = f"Inputs!$B${INPUT_ROW['Mid-year discounting (1/0)']}"
    TVMETH_CELL = f"Inputs!$B${INPUT_ROW['Exit Multiple (1/0)']}"
    EXITMULT_CELL = f"Inputs!$B${INPUT_ROW['Exit EV/EBITDA Multiple']}"
    DEBT_CELL   = f"Inputs!$B${INPUT_ROW['Total Debt']}"
    CASH_CELL   = f"Inputs!$B${INPUT_ROW['Cash & Equivalents']}"
    SHARES_CELL = f"Inputs!$B${INPUT_ROW['Shares Outstanding']}"
    PRICE_CELL  = f"Inputs!$B${INPUT_ROW['Current Stock Price']}"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12

    # =====================================================================
    # SHEET 2 — DCF Model (wide format: years as columns, formulas everywhere)
    # =====================================================================
    wsm = wb.create_sheet("DCF Model")
    wsm.sheet_view.showGridLines = False

    wsm["A1"] = "Discounted Cash Flow Model"
    wsm["A1"].font = f_title
    wsm["A2"] = f"{company_name} ({ticker})  |  All figures in $M unless noted  |  Every cell is a live formula"
    wsm["A2"].font = f_sub
    wsm.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + years)
    wsm.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + years)

    # Year headers (row 4): "Item" | Base | Year 1 | Year 2 | ... | Year N
    wsm.cell(row=4, column=1, value="Item").font = f_hdr
    wsm.cell(row=4, column=1).fill = fill_hdr; wsm.cell(row=4, column=1).alignment = center
    wsm.cell(row=4, column=1).border = box
    wsm.cell(row=4, column=2, value="Base (Yr 0)").font = f_hdr
    wsm.cell(row=4, column=2).fill = fill_hdr; wsm.cell(row=4, column=2).alignment = center
    wsm.cell(row=4, column=2).border = box
    for t in range(1, years + 1):
        c = wsm.cell(row=4, column=2 + t, value=f"Year {t}")
        c.font = f_hdr; c.fill = fill_hdr; c.alignment = center; c.border = box
    wsm.row_dimensions[4].height = 22

    # ── Revenue Build (uses per-year growth from path table on Inputs) ──
    wsm.cell(row=6, column=1, value="Revenue").font = f_bold
    wsm.cell(row=6, column=1).alignment = indent
    wsm.cell(row=6, column=2, value=f"={REV_CELL}").number_format = fmt_dollar
    for t in range(1, years + 1):
        col = 2 + t; prev = get_column_letter(col - 1)
        path_row = PATH_FIRST_ROW + (t - 1)
        wsm.cell(row=6, column=col,
            value=f"={prev}6*(1+Inputs!$B${path_row})").number_format = fmt_dollar

    # ── EBIT Margin (per-year from path table) ──
    wsm.cell(row=7, column=1, value="× EBIT Margin").alignment = indent
    wsm.cell(row=7, column=2, value="—").alignment = center
    for t in range(1, years + 1):
        col = 2 + t
        path_row = PATH_FIRST_ROW + (t - 1)
        wsm.cell(row=7, column=col, value=f"=Inputs!$C${path_row}").number_format = fmt_pct

    wsm.cell(row=8, column=1, value="= EBIT").font = f_bold
    wsm.cell(row=8, column=1).alignment = indent
    wsm.cell(row=8, column=2, value="—").alignment = center
    for t in range(1, years + 1):
        col = 2 + t; col_letter = get_column_letter(col)
        wsm.cell(row=8, column=col, value=f"={col_letter}6*{col_letter}7").number_format = fmt_dollar

    # ── NOPAT = EBIT × (1 - Tax) ──
    wsm.cell(row=9, column=1, value="× (1 − Tax Rate)").alignment = indent
    wsm.cell(row=9, column=2, value="—").alignment = center
    for t in range(1, years + 1):
        col = 2 + t
        wsm.cell(row=9, column=col, value=f"=1-{TAX_CELL}").number_format = fmt_pct

    wsm.cell(row=10, column=1, value="= NOPAT").font = f_bold
    wsm.cell(row=10, column=1).alignment = indent
    wsm.cell(row=10, column=2, value="—").alignment = center
    for t in range(1, years + 1):
        col = 2 + t; col_letter = get_column_letter(col)
        wsm.cell(row=10, column=col, value=f"={col_letter}8*{col_letter}9").number_format = fmt_dollar

    # ── Reinvestment ──
    wsm.cell(row=11, column=1, value="(−) Reinvestment").alignment = indent
    wsm.cell(row=11, column=2, value="—").alignment = center
    for t in range(1, years + 1):
        col = 2 + t; col_letter = get_column_letter(col)
        wsm.cell(row=11, column=col, value=f"=-{col_letter}10*{REINV_CELL}").number_format = fmt_dollar

    # ── ΔNWC drag (matches website's NWC line) ──
    wsm.cell(row=12, column=1, value="(−) ΔNWC").alignment = indent
    wsm.cell(row=12, column=2, value="—").alignment = center
    for t in range(1, years + 1):
        col = 2 + t; col_letter = get_column_letter(col); prev = get_column_letter(col - 1)
        wsm.cell(row=12, column=col,
            value=f"=-MAX(0,({col_letter}6-{prev}6)*{NWC_CELL})").number_format = fmt_dollar

    # ── FCF (highlighted) ──
    wsm.cell(row=13, column=1, value="= Free Cash Flow").font = f_summary
    wsm.cell(row=13, column=1).fill = fill_sum; wsm.cell(row=13, column=1).alignment = indent
    wsm.cell(row=13, column=1).border = high
    wsm.cell(row=13, column=2, value="—").alignment = center
    wsm.cell(row=13, column=2).fill = fill_sum; wsm.cell(row=13, column=2).border = high
    for t in range(1, years + 1):
        col = 2 + t; col_letter = get_column_letter(col)
        c = wsm.cell(row=13, column=col, value=f"={col_letter}10+{col_letter}11+{col_letter}12")
        c.number_format = fmt_dollar; c.font = f_summary
        c.fill = fill_sum; c.border = high

    # ── Discount Factor (mid-year aware via the 1/0 flag) & PV ──
    wsm.cell(row=15, column=1,
        value="Discount Factor: 1/(1+WACC)^(t − 0.5×midyr)").alignment = indent
    wsm.cell(row=15, column=2, value="—").alignment = center
    for t in range(1, years + 1):
        col = 2 + t
        wsm.cell(row=15, column=col,
            value=f"=1/(1+{WACC_CELL})^({t}-0.5*{MIDYR_CELL})").number_format = fmt_factor

    wsm.cell(row=16, column=1, value="PV of FCF").font = f_bold
    wsm.cell(row=16, column=1).alignment = indent
    wsm.cell(row=16, column=2, value="—").alignment = center
    for t in range(1, years + 1):
        col = 2 + t; col_letter = get_column_letter(col)
        wsm.cell(row=16, column=col, value=f"={col_letter}13*{col_letter}15").number_format = fmt_dollar

    # ── Valuation Summary ──
    last_yr_col = get_column_letter(2 + years)
    pv_first    = get_column_letter(3)
    pv_last     = get_column_letter(2 + years)

    summary = [
        ("Sum of PV of FCFs",       f"=SUM({pv_first}16:{pv_last}16)"),
        ("Terminal FCF (Yr N+1)",   f"={last_yr_col}13*(1+{TG_CELL})"),
        # IF Exit Multiple (TVMETH=1): exit_mult × terminal-year EBIT × 1.15 (EBITDA proxy)
        # ELSE Gordon Growth: terminal_FCF / (WACC − g)
        ("Terminal Value",
            f"=IF({TVMETH_CELL}=1,"
            f"{EXITMULT_CELL}*{last_yr_col}8*1.15,"
            f"B20/({WACC_CELL}-{TG_CELL}))"),
        ("PV of Terminal Value",
            f"=B21/(1+{WACC_CELL})^({YEARS_CELL}-0.5*{MIDYR_CELL})"),
        ("Enterprise Value",        f"=B19+B22"),
        ("(−) Total Debt",          f"=-{DEBT_CELL}"),
        ("(+) Cash & Equivalents",  f"={CASH_CELL}"),
        ("Equity Value",            f"=B23+B24+B25"),
        ("÷ Shares Outstanding",    f"={SHARES_CELL}"),
        ("Intrinsic Value / Share", f"=B26/B27"),
        ("Current Market Price",    f"={PRICE_CELL}"),
        ("Upside / (Downside)",     f"=B28/B29-1"),
    ]
    formats = [fmt_dollar, fmt_dollar, fmt_dollar, fmt_dollar, fmt_dollar,
               fmt_dollar, fmt_dollar, fmt_dollar, fmt_int, fmt_dol2, fmt_dol2, fmt_pct]

    # Section header for summary
    sh = wsm.cell(row=18, column=1, value="Valuation Summary")
    sh.font = f_hdr; sh.fill = fill_hdr; sh.alignment = indent
    wsm.merge_cells(start_row=18, start_column=1, end_row=18, end_column=2 + years)
    wsm.row_dimensions[18].height = 20

    for i, (label, formula) in enumerate(summary):
        r = 19 + i
        c1 = wsm.cell(row=r, column=1, value=label)
        c1.font = f_bold; c1.alignment = indent; c1.border = box
        c2 = wsm.cell(row=r, column=2, value=formula)
        c2.alignment = right; c2.border = box; c2.number_format = formats[i]
        if "Intrinsic Value" in label:
            c1.font = f_summary; c1.fill = fill_sum; c1.border = high
            c2.font = f_summary; c2.fill = fill_sum; c2.border = high

    # Column widths + freeze panes
    wsm.column_dimensions["A"].width = 30
    for t in range(years + 1):
        wsm.column_dimensions[get_column_letter(2 + t)].width = 13
    wsm.freeze_panes = "B5"

    return wb


col_dl, _ = st.columns([2, 5])
with col_dl:
    if st.button("📥 Generate Excel Workbook", type="primary", use_container_width=True):
        wb = build_excel_workbook()
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            label="⬇️ Download " + ticker + "_DCF_Replication.xlsx",
            data=buf.getvalue(),
            file_name=f"{ticker}_DCF_Replication.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.markdown("---")


# =============================================================================
# Methodology note (always visible at bottom)
# =============================================================================
with st.expander("📖 Methodology"):
    st.markdown("""
**Discounted Cash Flow Model**

Estimates a stock's intrinsic value from expected future cash flows, discounted to present value at the firm's cost of capital (WACC).

**Steps:**
1. **Forecast** revenue, EBIT, NOPAT, and free cash flow for an explicit period (3-15 years).
2. **Discount** each year's FCF to today using `1/(1+WACC)^t`.
3. **Terminal Value** captures everything beyond the forecast: `FCF_N × (1+g) / (WACC − g)`.
4. **Enterprise Value** = Σ PV of FCFs + PV of Terminal Value.
5. **Equity Value** = EV − Debt + Cash.
6. **Per-share** = Equity Value ÷ Diluted Shares Outstanding.

**WACC build-up (CAPM):** WACC = (E/V) × [Rf + β × ERP] + (D/V) × Rd × (1 − tax).

**Multi-stage growth (3-stage):** Years 1 to N₁ at the high-growth rate; Years N₁+1 to N₁+N₂ linearly fade to the terminal rate; Years beyond compound at the terminal rate (also used for the perpetuity).

**Margin expansion:** Optionally model margins improving (or compressing) linearly between Year 1 and Year N. Useful for growth firms with operating leverage.

**Input guidance:**
- **Growth rate** — informed by historical CAGR + forward thesis. Mature firms: a few % above GDP. Growth firms: double-digit, fading down.
- **EBIT margin** — historical operating margin. Better firms expand margins over time.
- **WACC** — riskier firms have higher WACC. Rough guide: large-cap stable 7–9%, mid-cap 9–12%, small/risky 12%+.
- **Terminal growth** — should not exceed long-run GDP (~2-3%) over the very long run.
- **Reinvestment rate** — % of NOPAT going back into CapEx + working capital. Higher reinvestment = less FCF today, more future growth.

**Data source:** SEC EDGAR Company Facts API (`data.sec.gov`) — official filings, used by professional analysts.
    """)

st.markdown("---")
st.caption(
    "⚠️ **For educational and informational purposes only — not investment advice.** "
    "DCF outputs are highly sensitive to assumptions. Results reflect your inputs, not guaranteed truth. "
    "Always do your own diligence before making investment decisions."
)
